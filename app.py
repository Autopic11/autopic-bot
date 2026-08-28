import os
from flask import Flask, request, jsonify
import requests
import openpyxl
from io import BytesIO

app = Flask(__name__)

CLIENT_ID = os.getenv("MS_CLIENT_ID", "14d82eec-204b-4c2f-b7e8-296a70dab67e")
REFRESH_TOKEN = os.getenv("MS_REFRESH_TOKEN")
FILE_PATH = os.getenv("ONEDRIVE_FILE_PATH", "AUTOPIC/DOCUMENTACIÓN/inventario.xlsx")
SHEET_NAME = os.getenv("SHEET_NAME", "Hoja2")

def get_access_token():
    resp = requests.post("https://login.microsoftonline.com/common/oauth2/v2.0/token", data={
        "client_id": CLIENT_ID,
        "grant_type": "refresh_token",
        "refresh_token": REFRESH_TOKEN,
        "scope": "Files.ReadWrite Files.ReadWrite.All offline_access User.Read"
    })
    data = resp.json()
    if "access_token" not in data:
        raise Exception(f"Error al refrescar token: {data}")
    return data["access_token"]

def download_excel(access_token):
    headers = {"Authorization": f"Bearer {access_token}"}
    # Buscar archivo por path
    url = f"https://graph.microsoft.com/v1.0/me/drive/root:/{FILE_PATH}:/content"
    r = requests.get(url, headers=headers)
    if r.status_code != 200:
        raise Exception(f"No se pudo descargar {FILE_PATH}: {r.text}")
    return r.content

def upload_excel(access_token, content_bytes):
    headers = {"Authorization": f"Bearer {access_token}", "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    url = f"https://graph.microsoft.com/v1.0/me/drive/root:/{FILE_PATH}:/content"
    r = requests.put(url, headers=headers, data=content_bytes)
    if r.status_code not in [200,201]:
        raise Exception(f"No se pudo subir: {r.text}")
    return r.json()

@app.route("/")
def home():
    return f"Autopic Bot corriendo! Archivo: {FILE_PATH} Hoja: {SHEET_NAME}"

@app.route('/agregar', methods=['GET', 'POST'])
def agregar():
    try:
        data = request.json
        # Esperado: {"producto":"Balata", "cantidad":4, "marca":"Brembo"}
        producto = data.get("producto", "")
        cantidad = data.get("cantidad", "")
        marca = data.get("marca", "")

        token = get_access_token()
        excel_bytes = download_excel(token)
        
        wb = openpyxl.load_workbook(BytesIO(excel_bytes))
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(["Producto", "Cantidad", "Marca", "Fecha"])
        else:
            ws = wb[SHEET_NAME]
        
        # Si está vacía, poner encabezado
        if ws.max_row == 1 and ws.cell(1,1).value is None:
            ws.append(["Producto", "Cantidad", "Marca", "Fecha"])
        
        from datetime import datetime
        ws.append([producto, cantidad, marca, datetime.now().strftime("%Y-%m-%d %H:%M")])
        
        output = BytesIO()
        wb.save(output)
        upload_excel(token, output.getvalue())

        return jsonify({"ok": True, "msg": f"Agregado a {FILE_PATH} en {SHEET_NAME}: {producto}"})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", 10000)))
